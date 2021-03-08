import pandas as pd
import datetime as dt
import time


# definition of a function to send an email if there is something to be consumed
def SendMailApp():
    import os
    import smtplib
    from email.message import EmailMessage

    EMAIL_ADRESS = os.environ.get('EM_USER')
    EMAIL_PASSWORD = os.environ.get('EM_PASSWORD')

    contacts = ['danilo.altomar@gmx.de']

    msg = EmailMessage()
    msg['Subject'] = 'Vencimento Alimentos'
    msg['From'] = EMAIL_ADRESS
    msg['To'] = contacts
    msg.set_content('Confira a lista para consumo')

    with open(excelfilename, 'rb') as file:
        file_data = file.read()
        file_name = file.name

    msg.add_attachment(file_data, maintype='application', subtype='octet-stream', filename=file_name)

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(EMAIL_ADRESS, EMAIL_PASSWORD)
        smtp.send_message(msg)


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


data = pd.read_excel("C:\\Users\\danil\\Desktop\\Prepper\\Lista_Mantimentos_Testes.xlsx", engine='xlrd')
today = dt.date.today()
control_date = today + dt.timedelta(60)
data['is_after_control_date'] = [x < control_date for x in data.Validade]
ready_to_consume = data[data['is_after_control_date'] == True]
append_df_to_excel('C:\\Users\\danil\\Desktop\\Prepper\\Lista_Consumo_Acumulada.xlsx', ready_to_consume, index=False, header=False)
indexTrue = data[data['is_after_control_date'] == True].index
data.drop(indexTrue, inplace=True)
data.drop(['is_after_control_date'], axis=1, inplace=True)
data.to_excel("C:\\Users\\danil\\Desktop\\Prepper\\Lista_Mantimentos_Testes.xlsx",index=False)

TodaysDate = time.strftime("%d-%m-%Y")
excelfilename = "C:\\Users\\danil\\Desktop\\Prepper\\"+TodaysDate + "_Lista_Consumo.xlsx"

if len(ready_to_consume)>0:
    ready_to_consume.to_excel(excelfilename, index=False)
    SendMailApp()


